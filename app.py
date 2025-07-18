import streamlit as st
# تصميم CSS مبسط يركز على الخط والاتجاه فقط
st.markdown("""
    <style>
    html, body, [class*="css"] {
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        direction: rtl;
        text-align: right;
    }
    h1 {
        color: #2c3e50;
    }
    </style>
""", unsafe_allow_html=True)


import pandas as pd
import re
from openpyxl import load_workbook
from hijri_converter import convert

# إعداد الصفحة
st.set_page_config(page_title="مساعد البحوث والدراسات", page_icon="📚", layout="wide")

# CSS لتعديل اتجاه النص ليكون من اليمين لليسار
st.markdown("""
<style>
    .main { direction: rtl; text-align: right; }
</style>
""", unsafe_allow_html=True)

st.title("📚 مساعد البحوث والدراسات")

st.markdown("""
**مرحبًا بك في مساعد البحوث والدراسات**

**هذه برمجيّة مُختصرة تعرض الجهد المميز للزملاء في ملف الإكسل (مجموع الدراسات) بطريقة أخرى**

يقتصر دور البرمجيّة على التعريف بالبحوث والدراسات دون تمكينها من الاطلاع على الملفات المرتبطة

البرمجيّة تعرض البحوث والدراسات حتى عام 1446هـ

**ابحث عن أي بحث أو دراسة بكلمة مفتاحيّة**

بالضغط مرتين على أي خانة تتّضح بشكل كامل

نسأل الله لنا ولكم التوفيق والسداد
""")

# تحميل البيانات
file_path = "مجموع الدراسات.xlsx"
wb = load_workbook(filename=file_path, data_only=True)
sheet = wb["ورقة1"]

# قراءة البيانات وتحويل التاريخ إلى هجري
data = []
for row in sheet.iter_rows():
    row_data = []
    for cell in row:
        if cell.is_date and cell.value:
            try:
                if 1900 <= cell.value.year <= 2100:
                    hijri_date = convert.Gregorian(cell.value.year, cell.value.month, cell.value.day).to_hijri()
                    row_data.append(f"{hijri_date.year}-{hijri_date.month}-{hijri_date.day}")
                else:
                    row_data.append(str(cell.value))
            except:
                row_data.append(str(cell.value))
        else:
            row_data.append(str(cell.value) if cell.value is not None else "")
    data.append(row_data)

# تحويلها إلى DataFrame
df = pd.DataFrame(data[1:], columns=data[0])

# حذف الأعمدة من موضوع 4 إلى موضوع 13
cols_to_drop = [col for col in df.columns if 'موضوع' in col and any(str(i) in col for i in range(4,14))]
df = df.drop(columns=cols_to_drop)

# حذف العمود الأول إذا لم يكن هو عمود (م)
if df.columns[0] != "م":
    df = df.drop(df.columns[0], axis=1)

def normalize_arabic(text):
    text = str(text)
    text = re.sub(r'[ًٌٍَُِّْـ]', '', text)  # حذف التشكيل
    text = re.sub(r'[إأآا]', 'ا', text)    # توحيد الألف
    text = re.sub(r'[ؤئ]', 'ء', text)      # توحيد الهمزة
    text = re.sub(r'ة', 'ه', text)         # تحويل التاء المربوطة
    text = re.sub(r'^ال', '', text)        # حذف أل التعريف من أول الكلمة
    return text



# واجهة البحث
st.markdown("---")
# تصميم صف يحتوي على مربع الإدخال وزر الإرسال
col1, col2 = st.columns([4,1])

with col1:
    study_name = st.text_input("🔎 تفضّل بالاستفسار عن أي دراسة:")

with col2:
    search_button = st.button("🔍 بحث")


if search_button and study_name:
    # تابع نفس منطق البحث السابق هنا...

    pattern = re.compile(normalize_arabic(study_name), re.IGNORECASE)
    filtered_data = df[df.apply(lambda row: row.astype(str).apply(lambda x: bool(pattern.search(normalize_arabic(str(x))))).any(), axis=1)]

    if not filtered_data.empty:
        st.success(f"✅ تم العثور على {len(filtered_data)} نتيجة مطابقة.")
        filtered_data = filtered_data.set_index("م")
        st.dataframe(filtered_data, use_container_width=True)
    else:
        st.warning("لم يتم العثور على أي نتائج تخصّ ما تفضّلتم به، لعل النتائج أن تظهر عند البحث بكلمة مفتاحيّة أخرى")
else:
    st.info("🔔 الرجاء كتابة كلمة مفتاحيّة ")
